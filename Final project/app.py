from __future__ import annotations

import csv
import io
import os
import random
import re
import smtplib
import base64
from collections import defaultdict
from datetime import datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from flask import (
    Flask,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from psycopg import connect as pg_connect
    from psycopg import OperationalError as PsycopgOperationalError
    from psycopg.rows import dict_row
    PSYCOPG_AVAILABLE = True
except ImportError:
    pg_connect = None
    dict_row = None
    PSYCOPG_AVAILABLE = False

    class PsycopgOperationalError(Exception):
        pass


# Core project paths and fixed defaults used throughout the portal.
BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "DAV_Project_csv.xlsx"
UPLOAD_DIR = BASE_DIR / "static" / "uploads"
EMAIL_SETTINGS_FILE = BASE_DIR / "email_settings.env"
SUPABASE_SETTINGS_FILE = BASE_DIR / "supabase_settings.env"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "Manoj@2005"
TABLE_NAME_MAP = {
    "users": "portal_users",
    "event_records": "event_results",
    "portal_events": "live_events",
    "event_competitions": "live_event_competitions",
    "event_registrations": "live_event_registrations",
    "activity_logs": "activity_history",
    "portal_settings": "website_settings",
    "schools": "school_list",
    "departments": "department_list",
}

app = Flask(__name__)
app.secret_key = "college-event-portal-secret"
DB_BACKEND_STATUS = {"engine": "postgres", "error": None}


# ---------------------------------------------------------------------------
# Settings and small utility helpers
# ---------------------------------------------------------------------------


def load_email_settings() -> None:
    # Load SMTP settings from a local env-style file for OTP email sending.
    if not EMAIL_SETTINGS_FILE.exists():
        return

    for line in EMAIL_SETTINGS_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_supabase_settings() -> None:
    # Load Supabase/Postgres settings without hard-coding secrets into the app.
    if not SUPABASE_SETTINGS_FILE.exists():
        return

    for line in SUPABASE_SETTINGS_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def now_ist() -> datetime:
    # Returns the current local time used for logs, OTP expiry, and timestamps.
    return datetime.now()


def now_iso() -> str:
    # Stores timestamps in a consistent string format for the database.
    return now_ist().isoformat(timespec="seconds")


def send_otp_email(recipient_email: str, recipient_name: str, otp_code: str) -> tuple[bool, str]:
    # Forgot-password OTP is sent by email through the configured SMTP account.
    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_port = int(os.getenv("SMTP_PORT", "587").strip() or "587")
    smtp_user = os.getenv("SMTP_USER", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip()
    smtp_sender = os.getenv("SMTP_SENDER", "").strip() or smtp_user

    if not all([smtp_host, smtp_port, smtp_user, smtp_password, smtp_sender]):
        return False, "SMTP email settings are missing."

    message = EmailMessage()
    message["Subject"] = "College Event Portal Password Reset OTP"
    message["From"] = smtp_sender
    message["To"] = recipient_email
    message.set_content(
        f"Hello {recipient_name},\n\n"
        f"Your OTP for College Event Portal password reset is: {otp_code}\n\n"
        "This OTP is valid for 5 minutes.\n"
        "If you did not request this, please ignore this email."
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(message)
        return True, "OTP sent to email successfully."
    except Exception as exc:
        return False, f"Unable to send OTP email: {exc}"


def to_int(value) -> int | None:
    # Safely converts different inputs into integers and returns None on bad values.
    try:
        if value in (None, "", "NA"):
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_date(value) -> datetime.date:
    # Normalizes both Excel datetime values and plain strings into a date object.
    if isinstance(value, datetime):
        return value.date()
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def is_postgres_enabled() -> bool:
    # True only when both the driver and the Supabase DB URL are available.
    return PSYCOPG_AVAILABLE and bool(os.getenv("SUPABASE_DB_URL", "").strip())


def normalize_postgres_url(url: str) -> str:
    # Supabase requires SSL, so this ensures sslmode is present in the connection string.
    normalized = url.strip()
    if normalized and "sslmode=" not in normalized:
        normalized = f"{normalized}?sslmode=require" if "?" not in normalized else f"{normalized}&sslmode=require"
    return normalized


def convert_query_for_postgres(query: str) -> str:
    # The app writes queries with ? placeholders and converts them for psycopg/Postgres.
    return query.replace("?", "%s")


def rewrite_query_table_names(query: str) -> str:
    # Keeps old logical table names in Python code while using clearer real table names in Supabase.
    rewritten = query
    for old_name, new_name in sorted(TABLE_NAME_MAP.items(), key=lambda item: len(item[0]), reverse=True):
        rewritten = re.sub(rf"\b{re.escape(old_name)}\b", new_name, rewritten)
    return rewritten


def rename_tables_for_clarity() -> None:
    # Renames existing old tables to the newer readable names once at startup.
    db = get_db()
    with db.cursor() as cursor:
        for old_name, new_name in TABLE_NAME_MAP.items():
            if old_name == new_name:
                continue
            cursor.execute(
                """
                SELECT EXISTS (
                    SELECT 1 FROM information_schema.tables
                    WHERE table_schema = 'public' AND table_name = %s
                ) AS present
                """,
                (old_name,),
            )
            old_exists = cursor.fetchone()["present"]
            cursor.execute(
                """
                SELECT EXISTS (
                    SELECT 1 FROM information_schema.tables
                    WHERE table_schema = 'public' AND table_name = %s
                ) AS present
                """,
                (new_name,),
            )
            new_exists = cursor.fetchone()["present"]
            if old_exists and not new_exists:
                cursor.execute(f'ALTER TABLE "{old_name}" RENAME TO "{new_name}"')
    db.commit()


# ---------------------------------------------------------------------------
# Database connection and query helpers
# ---------------------------------------------------------------------------


def get_db():
    # Opens one request-scoped database connection and reuses it through flask.g.
    if "db" not in g:
        if not PSYCOPG_AVAILABLE:
            raise RuntimeError("psycopg is required because this portal now uses Supabase/Postgres only.")
        db_url = os.getenv("SUPABASE_DB_URL", "").strip()
        if not db_url:
            raise RuntimeError("SUPABASE_DB_URL is missing. Configure Supabase before starting the portal.")
        try:
            connection = pg_connect(
                normalize_postgres_url(db_url),
                row_factory=dict_row,
                prepare_threshold=None,
            )
            DB_BACKEND_STATUS["engine"] = "postgres"
            DB_BACKEND_STATUS["error"] = None
        except PsycopgOperationalError as exc:
            DB_BACKEND_STATUS["engine"] = "postgres"
            DB_BACKEND_STATUS["error"] = str(exc)
            raise RuntimeError(f"Unable to connect to Supabase/Postgres: {exc}") from exc
        g.db = connection
    return g.db


@app.teardown_appcontext
def close_db(exception=None) -> None:
    # Closes the request-scoped connection after Flask finishes handling the request.
    connection = g.pop("db", None)
    if connection is not None:
        connection.close()


def fetch_one(query: str, params: tuple = ()):
    # Runs a SQL query and returns only the first matching row.
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(convert_query_for_postgres(rewrite_query_table_names(query)), params)
        return cursor.fetchone()


def fetch_all(query: str, params: tuple = ()) -> list:
    # Runs a SQL query and returns all matching rows.
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(convert_query_for_postgres(rewrite_query_table_names(query)), params)
        return cursor.fetchall()


def execute(query: str, params: tuple = ()) -> int:
    # Runs INSERT/UPDATE/DELETE queries and returns the new row id when possible.
    db = get_db()
    final_query = convert_query_for_postgres(rewrite_query_table_names(query))
    upper_query = final_query.lstrip().upper()
    if upper_query.startswith("INSERT INTO") and "RETURNING" not in upper_query:
        final_query = f"{final_query.rstrip().rstrip(';')} RETURNING id"
    with db.cursor() as cursor:
        cursor.execute(final_query, params)
        inserted = cursor.fetchone() if cursor.description else None
    db.commit()
    return inserted["id"] if inserted and isinstance(inserted, dict) and "id" in inserted else 0


def init_db() -> None:
    # Creates every table the portal needs if it does not exist yet in Supabase.
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            rewrite_query_table_names(
                """
        CREATE TABLE IF NOT EXISTS users (
            id BIGSERIAL PRIMARY KEY,
            role TEXT NOT NULL,
            username TEXT UNIQUE,
            password_hash TEXT,
            name TEXT NOT NULL,
            email TEXT UNIQUE,
            phone TEXT UNIQUE,
            registration_number TEXT UNIQUE,
            school_id BIGINT,
            school_name TEXT,
            department TEXT,
            study_year TEXT,
            last_login TEXT,
            otp_code TEXT,
            otp_created_at TEXT,
            profile_pic_path TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS event_records (
            id BIGSERIAL PRIMARY KEY,
            student_id TEXT,
            student_name TEXT NOT NULL,
            competition_name TEXT,
            school_name TEXT,
            department TEXT,
            study_year TEXT,
            event_id TEXT,
            event_name TEXT NOT NULL,
            event_date TEXT,
            team_id TEXT,
            team_size INTEGER,
            team_members TEXT,
            position TEXT,
            result TEXT,
            prize_money INTEGER DEFAULT 0,
            faculty_coordinator TEXT,
            venue TEXT,
            academic_year TEXT,
            created_by INTEGER,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS portal_events (
            id BIGSERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            total_prize_money INTEGER DEFAULT 0,
            competition_count INTEGER DEFAULT 0,
            event_year TEXT,
            school_id BIGINT,
            school_name TEXT,
            department_name TEXT,
            description TEXT,
            wallpaper_path TEXT,
            poster_image_path TEXT,
            registration_open BOOLEAN DEFAULT TRUE,
            created_by BIGINT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS event_competitions (
            id BIGSERIAL PRIMARY KEY,
            portal_event_id BIGINT NOT NULL,
            competition_name TEXT NOT NULL,
            venue TEXT,
            max_team_members INTEGER DEFAULT 1,
            first_prize INTEGER DEFAULT 0,
            second_prize INTEGER DEFAULT 0,
            third_prize INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS event_registrations (
            id BIGSERIAL PRIMARY KEY,
            portal_event_id BIGINT NOT NULL,
            user_id BIGINT NOT NULL,
            participant_name TEXT NOT NULL,
            participant_email TEXT,
            participant_phone TEXT,
            school_name TEXT,
            department TEXT,
            study_year TEXT,
            competition_id BIGINT,
            competition_name TEXT,
            team_name TEXT,
            team_id TEXT,
            team_members TEXT,
            notes TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS event_result_announcements (
            id BIGSERIAL PRIMARY KEY,
            portal_event_id BIGINT NOT NULL,
            competition_id BIGINT,
            competition_name TEXT,
            position TEXT,
            result_label TEXT,
            prize_money INTEGER DEFAULT 0,
            team_name TEXT,
            registration_number TEXT,
            participant_name TEXT NOT NULL,
            announced_by BIGINT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS activity_logs (
            id BIGSERIAL PRIMARY KEY,
            user_id BIGINT,
            action TEXT NOT NULL,
            name TEXT,
            email TEXT,
            phone TEXT,
            role TEXT,
            profile_pic_path TEXT,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS portal_settings (
            id BIGINT PRIMARY KEY,
            site_name TEXT DEFAULT 'College Event Portal',
            logo_path TEXT,
            wallpaper_path TEXT,
            login_wallpaper_path TEXT,
            dashboard_wallpaper_path TEXT,
            event_wallpaper_path TEXT,
            stats_wallpaper_path TEXT,
            winners_wallpaper_path TEXT,
            admin_wallpaper_path TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS schools (
            id BIGSERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS departments (
            id BIGSERIAL PRIMARY KEY,
            school_id BIGINT NOT NULL,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(school_id, name)
        );
        """
            )
        )
        cursor.execute(
            rewrite_query_table_names(
                "ALTER TABLE portal_events ADD COLUMN IF NOT EXISTS registration_open BOOLEAN DEFAULT TRUE"
            )
        )
        cursor.execute(
            rewrite_query_table_names(
                "ALTER TABLE event_competitions ADD COLUMN IF NOT EXISTS max_team_members INTEGER DEFAULT 1"
            )
        )
    db.commit()


# ---------------------------------------------------------------------------
# Startup seeding, migration, and shared data builders
# ---------------------------------------------------------------------------

def seed_admin() -> None:
    # Creates the default admin account only once on first startup.
    existing = fetch_one("SELECT id FROM users WHERE username = ?", (ADMIN_USERNAME,))
    if existing:
        return

    execute(
        """
        INSERT INTO users (role, username, password_hash, name, email, phone, department, study_year, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "admin",
            ADMIN_USERNAME,
            generate_password_hash(ADMIN_PASSWORD),
            "Manoj Admin",
            None,
            None,
            "Administration",
            "Admin",
            now_iso(),
        ),
    )


def import_workbook_if_empty() -> None:
    # Imports starter Excel data only when the records table is still empty.
    existing_count = fetch_one("SELECT COUNT(*) AS total FROM event_records")
    if existing_count and existing_count["total"] > 0:
        return

    if not DATA_FILE.exists():
        return

    workbook = load_workbook(DATA_FILE, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(max_row=1))]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        record = dict(zip(headers, row))
        event_date = parse_date(record.get("event_date")).isoformat()
        study_year = str(record.get("student_id", ""))[:2] if record.get("student_id") else ""
        execute(
            """
            INSERT INTO event_records (
                student_id, student_name, department, study_year, event_id, event_name, event_date,
                team_id, team_size, team_members, position, result, prize_money,
                faculty_coordinator, venue, academic_year, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(record.get("student_id") or ""),
                str(record.get("student_name") or ""),
                str(record.get("department") or ""),
                study_year,
                str(record.get("event_id") or ""),
                str(record.get("event_name") or ""),
                event_date,
                str(record.get("team_id") or ""),
                to_int(record.get("team_size")) or 0,
                str(record.get("team_members") or ""),
                str(record.get("position") or ""),
                str(record.get("result") or ""),
                to_int(record.get("prize_money")) or 0,
                str(record.get("faculty_coordinator") or ""),
                str(record.get("venue") or ""),
                event_date[:4],
                now_iso(),
            ),
        )


def init_portal() -> None:
    # App startup order matters here: create storage, bootstrap schema, migrate data, then seed defaults.
    # 1. Create folders/assets used by uploads and default profile pictures.
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    default_avatar = UPLOAD_DIR / "default_profile.svg"
    if not default_avatar.exists():
        default_avatar.write_text(
            """<svg xmlns="http://www.w3.org/2000/svg" width="160" height="160" viewBox="0 0 160 160"><rect width="160" height="160" rx="36" fill="#edf3ff"/><circle cx="80" cy="62" r="30" fill="#2555d9"/><path d="M31 139c8-28 26-42 49-42s41 14 49 42" fill="#5f83ff"/></svg>""",
            encoding="utf-8",
        )
    # 2. Rename older table names to the newer readable Supabase names if needed.
    rename_tables_for_clarity()
    # 3. Create tables and seed the first-use data needed by the portal.
    init_db()
    seed_admin()
    seed_schools_and_departments()
    import_workbook_if_empty()
    # 4. Normalize older imported rows so the live portal works with clean data.
    migrate_existing_data()
    # 5. Ensure one settings row exists for site name, logo, and wallpaper controls.
    if not fetch_one("SELECT id FROM portal_settings WHERE id = 1"):
        execute("INSERT INTO portal_settings (id, updated_at) VALUES (1, ?)", (now_iso(),))


@app.before_request
def load_current_user() -> None:
    # Pulls the logged-in user from the session before every request.
    user_id = session.get("user_id")
    g.current_user = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,)) if user_id else None


@app.context_processor
def inject_common_context() -> dict:
    # Shares common values with all templates so each route stays smaller.
    current_user = getattr(g, "current_user", None)
    portal_settings = fetch_one("SELECT * FROM portal_settings WHERE id = 1")
    schools = get_schools()
    departments = get_departments()
    endpoint = request.endpoint or ""
    page_wallpaper_map = {
        "home": "login_wallpaper_path",
        "signup": "login_wallpaper_path",
        "participant_login": "login_wallpaper_path",
        "forgot_password": "login_wallpaper_path",
        "admin_login": "login_wallpaper_path",
        "dashboard": "dashboard_wallpaper_path",
        "event_center": "event_wallpaper_path",
        "event_detail": "event_wallpaper_path",
        "events": "event_wallpaper_path",
        "stats_view": "stats_wallpaper_path",
        "leaderboard": "stats_wallpaper_path",
        "winners_view": "winners_wallpaper_path",
        "admin_records": "admin_wallpaper_path",
        "admin_accounts": "admin_wallpaper_path",
        "edit_record": "admin_wallpaper_path",
        "edit_account": "admin_wallpaper_path",
    }
    current_wallpaper = None
    if portal_settings:
        wallpaper_key = page_wallpaper_map.get(endpoint)
        if wallpaper_key:
            current_wallpaper = portal_settings[wallpaper_key]
        current_wallpaper = current_wallpaper or portal_settings["wallpaper_path"]
    return {
        "current_user": current_user,
        "logged_in": current_user is not None,
        "is_admin": bool(current_user and current_user["role"] == "admin"),
        "can_manage_events": bool(current_user and current_user["role"] in {"admin", "faculty", "event organiser"}),
        "can_access_event_manager": bool(current_user and current_user["role"] in {"admin", "faculty", "coordinator", "event organiser"}),
        "can_export": bool(current_user and current_user["role"] in {"admin", "coordinator", "event organiser"}),
        "portal_settings": portal_settings,
        "site_name": portal_settings["site_name"] if portal_settings and portal_settings["site_name"] else "College Event Portal",
        "current_wallpaper": current_wallpaper,
        "school_options": schools,
        "department_options_all": departments,
    }


def login_required(view):
    # Blocks access until a user is logged in.
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            flash("Please log in to access the portal.", "error")
            return redirect(url_for("home"))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    # Restricts access to the admin account only.
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            flash("Admin login is required.", "error")
            return redirect(url_for("admin_login"))
        if g.current_user["role"] != "admin":
            flash("Only admin can access that area.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


def event_manager_required(view):
    # Allows staff-style roles into the Event Manager.
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            flash("Login is required.", "error")
            return redirect(url_for("home"))
        if g.current_user["role"] not in {"admin", "faculty", "coordinator", "event organiser"}:
            flash("Only admin, faculty, coordinators, and event organisers can access event manager.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


def manage_events_required(view):
    # Allows only roles that are allowed to create or edit event content.
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            flash("Login is required.", "error")
            return redirect(url_for("home"))
        if g.current_user["role"] not in {"admin", "faculty", "event organiser"}:
            flash("Only admin, faculty, and event organisers can edit event data.", "error")
            return redirect(url_for("admin_records"))
        return view(*args, **kwargs)

    return wrapped


def export_required(view):
    # Restricts CSV downloads to approved roles.
    @wraps(view)
    def wrapped(*args, **kwargs):
        if g.current_user is None:
            flash("Login is required.", "error")
            return redirect(url_for("home"))
        if g.current_user["role"] not in {"admin", "coordinator", "event organiser"}:
            flash("Only admin, coordinators, and event organisers can export data.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)

    return wrapped


def update_last_login(user_id: int) -> None:
    # Stores the latest successful login time for profile/activity tracking.
    execute("UPDATE users SET last_login = ? WHERE id = ?", (now_iso(), user_id))


def render_chart_base64(plotter) -> str | None:
    # Renders a matplotlib chart into a base64 image string for direct display in templates.
    figure = None
    try:
        figure = plotter()
        if figure is None:
            return None
        buffer = io.BytesIO()
        figure.savefig(buffer, format="png", bbox_inches="tight", facecolor=figure.get_facecolor(), dpi=140)
        buffer.seek(0)
        return base64.b64encode(buffer.read()).decode("utf-8")
    finally:
        if figure is not None:
            plt.close(figure)


def build_dashboard_visuals(rows: list[dict]) -> list[dict]:
    # Builds the dashboard graph gallery from event and winner data using pandas/matplotlib.
    if not rows:
        return []

    df = pd.DataFrame(rows).copy()
    if df.empty:
        return []

    for column, fallback in {
        "school_name": "School of Engineering",
        "department": "General",
        "event_name": "Event",
        "competition_name": "",
        "academic_year": "",
        "position": "",
    }.items():
        if column not in df.columns:
            df[column] = fallback

    df["school_name"] = df["school_name"].fillna("School of Engineering").replace("", "School of Engineering")
    df["department"] = df["department"].fillna("General").replace("", "General")
    df["event_name"] = df["event_name"].fillna("Event").replace("", "Event")
    df["competition_name"] = df["competition_name"].fillna("")
    df.loc[df["competition_name"] == "", "competition_name"] = df.loc[df["competition_name"] == "", "event_name"]
    df["academic_year"] = df["academic_year"].fillna("").astype(str)
    df["position"] = df["position"].fillna("").astype(str).str.strip()
    winners_df = df[df["position"].isin(["1", "2", "3"])].copy()

    charts: list[dict] = []
    color_main = "#2555d9"
    color_alt = "#5f83ff"
    color_deep = "#173891"
    accent = "#ffb020"

    event_by_school = (
        df.groupby(["school_name", "event_name"]).size().reset_index(name="entries")
        .groupby("school_name")["event_name"].nunique().sort_values(ascending=False).head(8)
    )
    if not event_by_school.empty:
        def plot_event_by_school():
            fig, ax = plt.subplots(figsize=(6.4, 4.0), facecolor="#f8fbff")
            event_by_school.sort_values().plot(kind="barh", ax=ax, color=color_main)
            ax.set_title("Events by School", color="#13233d", fontsize=13, fontweight="bold")
            ax.set_xlabel("Unique events")
            ax.set_ylabel("")
            ax.grid(axis="x", alpha=0.2)
            ax.spines[["top", "right"]].set_visible(False)
            return fig
        charts.append({"title": "Events By School", "image": render_chart_base64(plot_event_by_school)})

    if not winners_df.empty:
        department_school = (
            df.groupby(["school_name", "department"]).size().reset_index(name="entries")
            .sort_values("entries", ascending=False)
        )
        top_departments = department_school.head(8)
        if not top_departments.empty:
            pivot = top_departments.pivot_table(index="department", columns="school_name", values="entries", fill_value=0)
            def plot_department_mix():
                fig, ax = plt.subplots(figsize=(7.0, 4.2), facecolor="#f8fbff")
                pivot.plot(kind="bar", stacked=True, ax=ax, color=[color_main, color_alt, accent, color_deep])
                ax.set_title("Department Entries by School", color="#13233d", fontsize=13, fontweight="bold")
                ax.set_xlabel("")
                ax.set_ylabel("Entries")
                ax.grid(axis="y", alpha=0.2)
                ax.spines[["top", "right"]].set_visible(False)
                ax.legend(title="School", fontsize=8, title_fontsize=9)
                return fig
            charts.append({"title": "Department Entries", "image": render_chart_base64(plot_department_mix)})

        competition_winners = winners_df.groupby("competition_name").size().sort_values(ascending=False).head(8)
        if not competition_winners.empty:
            def plot_competition_winners():
                fig, ax = plt.subplots(figsize=(5.8, 4.4), facecolor="#f8fbff")
                pie_data = competition_winners.head(6)
                ax.pie(
                    pie_data.values,
                    labels=pie_data.index,
                    autopct="%1.0f%%",
                    startangle=115,
                    colors=[accent, color_main, color_alt, color_deep, "#7dd3fc", "#60a5fa"],
                    textprops={"color": "#13233d", "fontsize": 8},
                )
                ax.set_title("Winners by Competition", color="#13233d", fontsize=13, fontweight="bold")
                return fig
            charts.append({"title": "Competition Winners", "image": render_chart_base64(plot_competition_winners)})

        yearly_winners = winners_df[winners_df["academic_year"] != ""].groupby("academic_year").size().sort_index()
        if not yearly_winners.empty:
            def plot_yearly_winners():
                fig, ax = plt.subplots(figsize=(6.4, 4.0), facecolor="#f8fbff")
                x_positions = list(range(len(yearly_winners.index)))
                ax.plot(x_positions, yearly_winners.values, marker="o", linewidth=2.5, color=color_deep)
                ax.fill_between(x_positions, yearly_winners.values, color=color_alt, alpha=0.18)
                ax.set_xticks(x_positions, yearly_winners.index)
                ax.set_title("Winners by Academic Year", color="#13233d", fontsize=13, fontweight="bold")
                ax.set_xlabel("Academic year")
                ax.set_ylabel("Winner count")
                ax.grid(alpha=0.2)
                ax.spines[["top", "right"]].set_visible(False)
                return fig
            charts.append({"title": "Winners By Year", "image": render_chart_base64(plot_yearly_winners)})

        winner_share = winners_df.groupby("school_name").size().sort_values(ascending=False).head(6)
        if not winner_share.empty:
            def plot_winner_share():
                fig, ax = plt.subplots(figsize=(6.6, 4.0), facecolor="#f8fbff")
                winner_share.sort_values().plot(kind="barh", ax=ax, color=color_main)
                ax.set_title("Winner Share by School", color="#13233d", fontsize=13, fontweight="bold")
                ax.set_xlabel("Winner count")
                ax.set_ylabel("")
                ax.grid(axis="x", alpha=0.2)
                ax.spines[["top", "right"]].set_visible(False)
                return fig
            charts.append({"title": "Winner Share By School", "image": render_chart_base64(plot_winner_share)})

    return [chart for chart in charts if chart.get("image")]


def filter_visual_rows(
    rows: list[dict],
    selected_school: str = "",
    selected_year: str = "",
    selected_event: str = "",
    selected_competition: str = "",
) -> list[dict]:
    # Applies the visuals page filters before building a selected chart.
    filtered = []
    for row in rows:
        if selected_school and (row.get("school_name") or "") != selected_school:
            continue
        if selected_year and str(row.get("academic_year") or "") != selected_year:
            continue
        if selected_event and (row.get("event_name") or "") != selected_event:
            continue
        row_competition = row.get("competition_name") or row.get("event_name") or ""
        if selected_competition and row_competition != selected_competition:
            continue
        filtered.append(row)
    return filtered


def build_visual_rows(records: list[dict], created_events: list[dict]) -> list[dict]:
    # Builds a richer visuals dataset so newly created live events appear even before historic rows exist.
    visual_rows = [dict(row) for row in records]
    seen_keys = {
        (
            row.get("event_name") or "",
            row.get("competition_name") or row.get("event_name") or "",
            row.get("school_name") or "",
            str(row.get("academic_year") or ""),
            row.get("department") or "",
        )
        for row in visual_rows
    }

    for event in created_events:
        event_name = event.get("title") or ""
        school_name = event.get("school_name") or ""
        department_name = event.get("department_name") or ""
        academic_year = str(event.get("event_year") or "")
        competitions = event.get("competitions") or []

        if not competitions:
            key = (event_name, event_name, school_name, academic_year, department_name)
            if key not in seen_keys:
                visual_rows.append(
                    {
                        "event_name": event_name,
                        "competition_name": event_name,
                        "school_name": school_name,
                        "department": department_name,
                        "academic_year": academic_year,
                        "student_name": "",
                        "position": "",
                        "result": "",
                        "prize_money": to_int(event.get("total_prize_money")) or 0,
                    }
                )
                seen_keys.add(key)
            continue

        for competition in competitions:
            competition_name = competition.get("competition_name") or event_name
            key = (event_name, competition_name, school_name, academic_year, department_name)
            if key not in seen_keys:
                visual_rows.append(
                    {
                        "event_name": event_name,
                        "competition_name": competition_name,
                        "school_name": school_name,
                        "department": department_name,
                        "academic_year": academic_year,
                        "student_name": "",
                        "position": "",
                        "result": "",
                        "prize_money": max(
                            to_int(competition.get("first_prize")) or 0,
                            to_int(competition.get("second_prize")) or 0,
                            to_int(competition.get("third_prize")) or 0,
                        ),
                    }
                )
                seen_keys.add(key)
    return visual_rows


def build_selected_visual(rows: list[dict], graph_key: str, graph_type: str = "auto") -> dict | None:
    # Builds one chosen chart for the dedicated visuals page.
    if not rows:
        return None

    df = pd.DataFrame(rows).copy()
    if df.empty:
        return None

    for column, fallback in {
        "school_name": "School of Engineering",
        "department": "General",
        "event_name": "Event",
        "competition_name": "",
        "academic_year": "",
        "position": "",
    }.items():
        if column not in df.columns:
            df[column] = fallback

    df["school_name"] = df["school_name"].fillna("School of Engineering").replace("", "School of Engineering")
    df["department"] = df["department"].fillna("General").replace("", "General")
    df["event_name"] = df["event_name"].fillna("Event").replace("", "Event")
    df["competition_name"] = df["competition_name"].fillna("")
    df.loc[df["competition_name"] == "", "competition_name"] = df.loc[df["competition_name"] == "", "event_name"]
    df["academic_year"] = df["academic_year"].fillna("").astype(str)
    df["position"] = df["position"].fillna("").astype(str).str.strip()
    winners_df = df[df["position"].isin(["1", "2", "3"])].copy()
    color_main = "#2555d9"
    color_alt = "#5f83ff"
    color_deep = "#173891"
    accent = "#ffb020"
    graph_type = (graph_type or "auto").strip().lower()

    if graph_key == "events_by_school":
        series = df.groupby("school_name")["event_name"].nunique().sort_values(ascending=False).head(8)
        if series.empty:
            return None

        def plotter():
            fig, ax = plt.subplots(figsize=(5.6, 3.6), facecolor="#f8fbff")
            if graph_type == "pie":
                ax.pie(
                    series.values,
                    labels=series.index,
                    autopct="%1.0f%%",
                    startangle=115,
                    colors=[color_main, color_alt, accent, color_deep, "#7dd3fc", "#60a5fa"],
                    textprops={"color": "#13233d", "fontsize": 8},
                )
            elif graph_type == "line":
                x_positions = list(range(len(series.index)))
                ax.plot(x_positions, series.values, marker="o", linewidth=2.5, color=color_deep)
                ax.fill_between(x_positions, series.values, color=color_alt, alpha=0.18)
                ax.set_xticks(x_positions, series.index, rotation=20, ha="right")
                ax.set_ylabel("Unique events")
            else:
                series.sort_values().plot(kind="barh", ax=ax, color=color_main)
                ax.set_xlabel("Unique events")
                ax.set_ylabel("")
                ax.grid(axis="x", alpha=0.2)
            ax.set_title("Events by School", color="#13233d", fontsize=13, fontweight="bold")
            ax.spines[["top", "right"]].set_visible(False)
            return fig

        return {"title": "Events By School", "image": render_chart_base64(plotter)}

    if graph_key == "department_entries":
        pivot = (
            df.groupby(["department", "school_name"]).size().reset_index(name="entries")
            .pivot_table(index="department", columns="school_name", values="entries", fill_value=0)
        )
        pivot = pivot.loc[pivot.sum(axis=1).sort_values(ascending=False).index]
        if pivot.empty:
            return None
        department_totals = pivot.sum(axis=1).head(8)

        def plotter():
            fig, ax = plt.subplots(figsize=(5.8, 3.8), facecolor="#f8fbff")
            if graph_type == "pie":
                ax.pie(
                    department_totals.values,
                    labels=department_totals.index,
                    autopct="%1.0f%%",
                    startangle=115,
                    colors=[color_main, color_alt, accent, color_deep, "#7dd3fc", "#60a5fa", "#1d4ed8", "#38bdf8"],
                    textprops={"color": "#13233d", "fontsize": 8},
                )
            elif graph_type == "line":
                x_positions = list(range(len(department_totals.index)))
                ax.plot(x_positions, department_totals.values, marker="o", linewidth=2.5, color=color_deep)
                ax.set_xticks(x_positions, department_totals.index, rotation=20, ha="right")
                ax.set_ylabel("Entries")
                ax.grid(alpha=0.2)
            else:
                pivot.head(8).plot(kind="bar", stacked=True, ax=ax, color=[color_main, color_alt, accent, color_deep])
                ax.set_ylabel("Entries")
                ax.grid(axis="y", alpha=0.2)
                ax.legend(title="School", fontsize=8, title_fontsize=9)
            ax.set_title("Department Entries by School", color="#13233d", fontsize=13, fontweight="bold")
            ax.set_xlabel("")
            ax.spines[["top", "right"]].set_visible(False)
            return fig

        return {"title": "Department Entries", "image": render_chart_base64(plotter)}

    if graph_key == "competition_winners":
        series = winners_df.groupby("competition_name").size().sort_values(ascending=False).head(6)
        if series.empty:
            return None

        def plotter():
            fig, ax = plt.subplots(figsize=(5.8, 3.8), facecolor="#f8fbff")
            if graph_type == "bar":
                series.plot(kind="bar", ax=ax, color=[accent, color_main, color_alt, color_deep, "#7dd3fc", "#60a5fa"])
                ax.set_ylabel("Winner count")
                ax.set_xlabel("")
                ax.grid(axis="y", alpha=0.2)
                ax.tick_params(axis="x", rotation=18)
            elif graph_type == "line":
                x_positions = list(range(len(series.index)))
                ax.plot(x_positions, series.values, marker="o", linewidth=2.5, color=color_deep)
                ax.set_xticks(x_positions, series.index, rotation=20, ha="right")
                ax.set_ylabel("Winner count")
                ax.grid(alpha=0.2)
            else:
                ax.pie(
                    series.values,
                    labels=series.index,
                    autopct="%1.0f%%",
                    startangle=115,
                    colors=[accent, color_main, color_alt, color_deep, "#7dd3fc", "#60a5fa"],
                    textprops={"color": "#13233d", "fontsize": 8},
                )
            ax.set_title("Winners by Competition", color="#13233d", fontsize=13, fontweight="bold")
            return fig

        return {"title": "Competition Winners", "image": render_chart_base64(plotter)}

    if graph_key == "winners_by_year":
        series = winners_df[winners_df["academic_year"] != ""].groupby("academic_year").size().sort_index()
        if series.empty:
            return None

        def plotter():
            fig, ax = plt.subplots(figsize=(5.8, 3.6), facecolor="#f8fbff")
            if graph_type == "bar":
                series.plot(kind="bar", ax=ax, color=color_main)
                ax.set_ylabel("Winner count")
                ax.grid(axis="y", alpha=0.2)
            elif graph_type == "pie":
                ax.pie(
                    series.values,
                    labels=series.index,
                    autopct="%1.0f%%",
                    startangle=115,
                    colors=[color_main, color_alt, accent, color_deep, "#7dd3fc", "#60a5fa"],
                    textprops={"color": "#13233d", "fontsize": 8},
                )
            else:
                x_positions = list(range(len(series.index)))
                ax.plot(x_positions, series.values, marker="o", linewidth=2.5, color=color_deep)
                ax.fill_between(x_positions, series.values, color=color_alt, alpha=0.18)
                ax.set_xticks(x_positions, series.index)
                ax.set_ylabel("Winner count")
                ax.grid(alpha=0.2)
            ax.set_title("Winners by Academic Year", color="#13233d", fontsize=13, fontweight="bold")
            ax.set_xlabel("Academic year")
            ax.spines[["top", "right"]].set_visible(False)
            return fig

        return {"title": "Winners By Year", "image": render_chart_base64(plotter)}

    if graph_key == "winner_share_school":
        series = winners_df.groupby("school_name").size().sort_values(ascending=False).head(8)
        if series.empty:
            return None

        def plotter():
            fig, ax = plt.subplots(figsize=(5.8, 3.6), facecolor="#f8fbff")
            if graph_type == "pie":
                ax.pie(
                    series.values,
                    labels=series.index,
                    autopct="%1.0f%%",
                    startangle=115,
                    colors=[color_main, color_alt, accent, color_deep, "#7dd3fc", "#60a5fa"],
                    textprops={"color": "#13233d", "fontsize": 8},
                )
            elif graph_type == "line":
                x_positions = list(range(len(series.index)))
                ax.plot(x_positions, series.values, marker="o", linewidth=2.5, color=color_deep)
                ax.set_xticks(x_positions, series.index, rotation=20, ha="right")
                ax.set_ylabel("Winner count")
                ax.grid(alpha=0.2)
            else:
                series.sort_values().plot(kind="barh", ax=ax, color=color_main)
                ax.set_xlabel("Winner count")
                ax.set_ylabel("")
                ax.grid(axis="x", alpha=0.2)
            ax.set_title("Winner Share by School", color="#13233d", fontsize=13, fontweight="bold")
            ax.spines[["top", "right"]].set_visible(False)
            return fig

        return {"title": "Winner Share By School", "image": render_chart_base64(plotter)}

    return None


def get_announcement_rows() -> list[dict]:
    # Converts live announced results into winner-like rows for pages that need a shared result view.
    rows = fetch_all(
        """
        SELECT
            a.*,
            e.title AS event_name,
            e.school_name AS event_school_name,
            e.department_name AS event_department_name,
            e.event_year AS event_year,
            u.name AS user_name,
            u.school_name AS user_school_name,
            u.department AS user_department
        FROM event_result_announcements a
        JOIN portal_events e ON e.id = a.portal_event_id
        LEFT JOIN users u ON upper(u.registration_number) = upper(a.registration_number)
        ORDER BY a.created_at DESC, a.id DESC
        """
    )
    result_rows = []
    for row in rows:
        data = dict(row)
        result_rows.append(
            {
                "id": data["id"],
                "event_name": data.get("event_name") or "Event",
                "competition_name": data.get("competition_name") or data.get("event_name") or "Competition",
                "school_name": data.get("user_school_name") or data.get("event_school_name") or "",
                "department": data.get("user_department") or data.get("event_department_name") or "",
                "student_name": data.get("participant_name") or data.get("user_name") or "Participant",
                "position": str(data.get("position") or "").strip(),
                "result": data.get("result_label") or "",
                "prize_money": data.get("prize_money") or 0,
                "academic_year": str(data.get("event_year") or ""),
                "registration_number": data.get("registration_number") or "",
                "source": "announcement",
            }
        )
    return result_rows


def get_user_history(account: dict) -> dict:
    # Collects profile-specific registration and result history for one account.
    registration_number = (account.get("registration_number") or "").strip().upper()
    registration_history = [
        dict(row)
        for row in fetch_all(
            """
            SELECT
                r.*,
                e.title AS event_title,
                e.event_year,
                e.school_name AS event_school_name,
                e.department_name AS event_department_name
            FROM event_registrations r
            JOIN portal_events e ON e.id = r.portal_event_id
            WHERE r.user_id = ?
            ORDER BY r.created_at DESC, r.id DESC
            """,
            (account["id"],),
        )
    ]
    record_history = [
        dict(row)
        for row in fetch_all(
            """
            SELECT * FROM event_records
            WHERE lower(student_name) = lower(?)
               OR student_id = ?
            ORDER BY event_date DESC, id DESC
            """,
            (account["name"], registration_number),
        )
    ]
    announced_wins = []
    if registration_number:
        announced_wins = [
            dict(row)
            for row in fetch_all(
                """
                SELECT
                    a.*,
                    e.title AS event_name,
                    e.event_year,
                    e.school_name AS event_school_name,
                    e.department_name AS event_department_name
                FROM event_result_announcements a
                JOIN portal_events e ON e.id = a.portal_event_id
                WHERE upper(a.registration_number) = ?
                ORDER BY a.created_at DESC, a.id DESC
                """,
                (registration_number,),
            )
        ]
    combined_history = [
        {
            "event_name": row.get("event_name") or "",
            "competition_name": row.get("competition_name") or row.get("event_name") or "",
            "academic_year": row.get("academic_year") or "",
            "result": row.get("result") or "",
            "prize_money": row.get("prize_money") or 0,
            "source": "record",
            "sort_date": str(row.get("event_date") or ""),
        }
        for row in record_history
    ]
    combined_history.extend(
        {
            "event_name": row.get("event_name") or "",
            "competition_name": row.get("competition_name") or row.get("event_name") or "",
            "academic_year": row.get("event_year") or "",
            "result": row.get("result_label") or row.get("position") or "",
            "prize_money": row.get("prize_money") or 0,
            "source": "announcement",
            "sort_date": str(row.get("created_at") or ""),
        }
        for row in announced_wins
    )
    combined_history = sorted(
        combined_history,
        key=lambda row: (
            str(row.get("sort_date") or ""),
            str(row.get("event_name") or "").lower(),
            str(row.get("competition_name") or "").lower(),
        ),
        reverse=True,
    )
    return {
        "registration_history": registration_history,
        "record_history": record_history,
        "announced_wins": announced_wins,
        "combined_history": combined_history,
    }


def build_portal_data() -> dict:
    # Build the shared dashboard/statistics payload once so multiple pages can reuse it.
    # This is the main summary builder for dashboard, stats, winners, and reports.
    rows = [dict(row) for row in fetch_all("SELECT * FROM event_records ORDER BY event_date DESC, id DESC")]
    events_map: dict[str, dict] = {}
    students: dict[str, dict] = {}
    leaderboard: dict[str, dict] = defaultdict(
        lambda: {"school": "", "points": 0, "wins": 0, "prize_money": 0, "participants": set()}
    )
    recent_results = []

    for row in rows:
        # Normalize each record into the derived structures used by the UI.
        event_date_value = row.get("event_date") or now_ist().date().isoformat()
        parsed_date = parse_date(event_date_value)
        row["event_date_display"] = parsed_date.strftime("%d %b %Y")
        event_id = row.get("event_id") or f"EV-{row['id']}"
        department = row.get("department") or "General"
        school_name = row.get("school_name") or "School of Engineering"
        competition_name = row.get("competition_name") or row.get("event_name")
        student_key = row.get("student_id") or f"guest-{row['id']}"

        event = events_map.setdefault(
            event_id,
            {
                "event_id": event_id,
                "event_name": row["event_name"],
                "school_name": school_name,
                "event_date_display": row["event_date_display"],
                "event_date_obj": parsed_date,
                "faculty_coordinator": row.get("faculty_coordinator") or "Not assigned",
                "venue": row.get("venue") or "Campus",
                "participants": set(),
                "departments": set(),
                "teams": set(),
                "prize_pool": 0,
                "winner_names": [],
                "result_labels": set(),
                "academic_year": row.get("academic_year") or parsed_date.strftime("%Y"),
            },
        )
        event["participants"].add(student_key)
        event["departments"].add(department)
        if row.get("team_id"):
            event["teams"].add(row["team_id"])
        event["prize_pool"] = max(event["prize_pool"], to_int(row.get("prize_money")) or 0)
        if row.get("result"):
            event["result_labels"].add(row["result"])
        if str(row.get("position") or "").strip() == "1":
            event["winner_names"].append(row["student_name"])

        students[student_key] = {
            "student_id": row.get("student_id") or "N/A",
            "student_name": row["student_name"],
            "school_name": school_name,
            "department": department,
            "study_year": row.get("study_year") or "-",
            "events_joined": students.get(student_key, {}).get("events_joined", 0) + 1,
            "best_result": row.get("result") or "Participant",
        }

        # Rankings are now grouped by school instead of department.
        school_board = leaderboard[school_name]
        school_board["school"] = school_name
        school_board["participants"].add(student_key)
        school_board["prize_money"] += to_int(row.get("prize_money")) or 0
        position = str(row.get("position") or "").strip()
        if position == "1":
            school_board["wins"] += 1
            school_board["points"] += 10
        elif position == "2":
            school_board["points"] += 6
        elif position == "3":
            school_board["points"] += 4
        else:
            school_board["points"] += 1

        if row.get("result") in {"Winner", "Runner-up", "Second Runner-up"}:
            recent_results.append(
                {
                    "event_name": row["event_name"],
                    "competition_name": competition_name,
                    "student_name": row["student_name"],
                    "school_name": school_name,
                    "department": department,
                    "result": row["result"],
                    "prize_money": to_int(row.get("prize_money")) or 0,
                    "venue": row.get("venue") or "Campus",
                }
            )

    events = sorted(events_map.values(), key=lambda item: item["event_date_obj"], reverse=True)
    for event in events:
        event["participant_count"] = len(event["participants"])
        event["department_count"] = len(event["departments"])
        event["team_count"] = len(event["teams"])
        event["winner_names"] = ", ".join(sorted(set(event["winner_names"]))[:4]) if event["winner_names"] else "TBD"
        event["result_summary"] = ", ".join(sorted(event["result_labels"])) if event["result_labels"] else "Open"
        event["departments"] = sorted(event["departments"])

    department_rankings = sorted(
        (
            {
                "school": item["school"],
                "points": item["points"],
                "wins": item["wins"],
                "prize_money": item["prize_money"],
                "participant_count": len(item["participants"]),
            }
            for item in leaderboard.values()
        ),
        key=lambda item: (-item["points"], -item["wins"], item["school"]),
    )
    top_students = sorted(students.values(), key=lambda item: (-item["events_joined"], item["student_name"]))[:10]
    registered_users = fetch_one("SELECT COUNT(*) AS total FROM users WHERE role != 'admin'")
    distinct_years = fetch_all("SELECT DISTINCT academic_year FROM event_records WHERE academic_year IS NOT NULL AND academic_year != '' ORDER BY academic_year DESC")
    created_events = fetch_all("SELECT * FROM portal_events ORDER BY created_at DESC")
    dashboard_visuals = build_dashboard_visuals(rows)

    return {
        "records": rows,
        "events": events,
        "department_rankings": department_rankings,
        "recent_results": sorted(recent_results, key=lambda item: (-item["prize_money"], item["event_name"]))[:8],
        "top_students": top_students,
        "available_years": [row["academic_year"] for row in distinct_years],
        "created_events": created_events,
        "dashboard_visuals": dashboard_visuals,
        "stats": {
            "events": len(events),
            "participants": len(students),
            "departments": len(department_rankings),
            "schools": len({row.get("school_name") or "School of Engineering" for row in rows}),
            "prize_pool": sum(item["prize_pool"] for item in events),
            "registered_users": registered_users["total"] if registered_users else 0,
        },
    }


def get_created_events_with_competitions() -> list[dict]:
    # Loads live events and attaches their competitions and registration counts.
    events = [dict(row) for row in fetch_all("SELECT * FROM portal_events ORDER BY created_at DESC")]
    for event in events:
        event["competitions"] = [
            dict(row)
            for row in fetch_all(
                "SELECT * FROM event_competitions WHERE portal_event_id = ? ORDER BY id",
                (event["id"],),
            )
        ]
        event["registration_count"] = fetch_one(
            "SELECT COUNT(*) AS total FROM event_registrations WHERE portal_event_id = ?",
            (event["id"],),
        )["total"]
    return events


def get_event_by_id(event_id: int) -> dict | None:
    # Loads one live event along with its competitions and registrations.
    event = fetch_one("SELECT * FROM portal_events WHERE id = ?", (event_id,))
    if not event:
        return None
    data = dict(event)
    data["registration_open"] = bool(data.get("registration_open", True))
    data["competitions"] = [
        dict(row)
        for row in fetch_all(
            "SELECT * FROM event_competitions WHERE portal_event_id = ? ORDER BY id",
            (event_id,),
        )
    ]
    data["registrations"] = [
        dict(row)
        for row in fetch_all(
            """
            SELECT
                r.*,
                u.registration_number,
                u.name AS account_name
            FROM event_registrations r
            LEFT JOIN users u ON u.id = r.user_id
            WHERE r.portal_event_id = ?
            ORDER BY r.created_at DESC
            """,
            (event_id,),
        )
    ]
    data["announcements"] = [
        dict(row)
        for row in fetch_all(
            "SELECT * FROM event_result_announcements WHERE portal_event_id = ? ORDER BY created_at DESC, id DESC",
            (event_id,),
        )
    ]
    data["announcement_count"] = len(data["announcements"])
    return data


def save_uploaded_file(file_storage, prefix: str) -> str | None:
    # Saves uploaded files into static/uploads and returns the relative public path.
    if not file_storage or not file_storage.filename:
        return None
    filename = secure_filename(file_storage.filename)
    if not filename:
        return None
    final_name = f"{prefix}_{int(datetime.now().timestamp())}_{filename}"
    path = UPLOAD_DIR / final_name
    file_storage.save(path)
    return f"uploads/{final_name}"


def default_profile_pic() -> str:
    # Default avatar used when a user does not upload a profile picture.
    return "uploads/default_profile.svg"


def get_schools() -> list[dict]:
    # Returns all schools for dropdowns and admin pages.
    return [dict(row) for row in fetch_all("SELECT * FROM schools ORDER BY name")]


def get_departments() -> list[dict]:
    # Returns all departments across every school.
    return [dict(row) for row in fetch_all("SELECT * FROM departments ORDER BY name")]


def get_departments_for_school(school_id: int | None) -> list[dict]:
    # Returns departments only for the selected school.
    if not school_id:
        return []
    return [dict(row) for row in fetch_all("SELECT * FROM departments WHERE school_id = ? ORDER BY name", (school_id,))]


def get_school_by_id(school_id: int | None):
    # Converts a school id into the full school row.
    if not school_id:
        return None
    return fetch_one("SELECT * FROM schools WHERE id = ?", (school_id,))


def seed_schools_and_departments() -> None:
    # Seeds the default school/department structure used by the portal.
    if not fetch_one("SELECT id FROM schools LIMIT 1"):
        school_names = ["School of Engineering", "School of Biosciences"]
        for school_name in school_names:
            execute("INSERT INTO schools (name, created_at) VALUES (?, ?)", (school_name, now_iso()))

    engineering = fetch_one("SELECT id FROM schools WHERE name = ?", ("School of Engineering",))
    biosciences = fetch_one("SELECT id FROM schools WHERE name = ?", ("School of Biosciences",))
    default_departments = []
    if engineering:
        default_departments.extend(
            [
                (engineering["id"], "Computer Science and Engineering"),
                (engineering["id"], "Electronics and Communication Engineering"),
                (engineering["id"], "Artificial Intelligence"),
            ]
        )
    if biosciences:
        default_departments.extend(
            [
                (biosciences["id"], "Biotechnology"),
                (biosciences["id"], "Microbiology"),
                (biosciences["id"], "Biochemistry"),
            ]
        )
    for school_id, department_name in default_departments:
        if not fetch_one("SELECT id FROM departments WHERE school_id = ? AND name = ?", (school_id, department_name)):
            execute("INSERT INTO departments (school_id, name, created_at) VALUES (?, ?, ?)", (school_id, department_name, now_iso()))


def migrate_existing_data() -> None:
    # Normalizes older rows so historic data fits the newer live-portal structure.
    samyuti_rows = fetch_all("SELECT id, event_name FROM event_records WHERE competition_name IS NULL OR competition_name = ''")
    for row in samyuti_rows:
        execute(
            "UPDATE event_records SET competition_name = ?, event_name = ?, event_id = COALESCE(NULLIF(event_id, ''), 'SAMYUTI-2026') WHERE id = ?",
            (row["event_name"], "Samyuti", row["id"]),
        )

    execute(
        "UPDATE event_records SET school_name = COALESCE(NULLIF(school_name, ''), 'School of Engineering') WHERE school_name IS NULL OR school_name = ''"
    )
    execute("UPDATE event_records SET school_name = 'School of Engineering' WHERE event_name = 'Samyuti'")
    execute("UPDATE event_records SET event_id = 'SAMYUTI-2026' WHERE event_name = 'Samyuti'")
    execute("UPDATE event_records SET event_date = COALESCE(NULLIF(event_date, ''), '2026-03-28') WHERE event_name = 'Samyuti'")
    execute(
        "UPDATE users SET school_name = COALESCE(NULLIF(school_name, ''), 'School of Engineering') WHERE role != 'admin' AND (school_name IS NULL OR school_name = '')"
    )

    for school in get_schools():
        execute(
            "UPDATE users SET school_id = ? WHERE school_name = ? AND school_id IS NULL",
            (school["id"], school["name"]),
        )

    engineering = fetch_one("SELECT id FROM schools WHERE name = ?", ("School of Engineering",))
    if engineering:
        samyuti_event = fetch_one("SELECT id FROM portal_events WHERE title = 'Samyuti'")
        if samyuti_event:
            execute(
                """
                UPDATE portal_events
                SET school_id = ?, school_name = ?, department_name = COALESCE(NULLIF(department_name, ''), 'School of Engineering'),
                    event_year = COALESCE(NULLIF(event_year, ''), '2026'),
                    description = COALESCE(NULLIF(description, ''), 'Many streams. One celebration.')
                WHERE id = ?
                """,
                (engineering["id"], "School of Engineering", samyuti_event["id"]),
            )
        elif fetch_one("SELECT id FROM event_records WHERE event_name = 'Samyuti'"):
            execute(
                """
                INSERT INTO portal_events (
                    title, total_prize_money, competition_count, event_year, school_id, school_name, department_name,
                    description, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "Samyuti",
                    0,
                    len(
                        {
                            row["competition_name"]
                            for row in fetch_all("SELECT competition_name FROM event_records WHERE event_name = 'Samyuti'")
                            if row["competition_name"]
                        }
                    ),
                    "2026",
                    engineering["id"],
                    "School of Engineering",
                    "School of Engineering",
                    "Many streams. One celebration.",
                    1,
                    now_iso(),
                ),
            )


def log_activity(action: str, user: dict | None = None) -> None:
    # Writes a lightweight activity entry so admin can see recent account actions.
    if user is None:
        user = getattr(g, "current_user", None)
    if user is None:
        execute(
            "INSERT INTO activity_logs (action, created_at) VALUES (?, ?)",
            (action, now_iso()),
        )
        return

    execute(
        """
        INSERT INTO activity_logs (user_id, action, name, email, phone, role, profile_pic_path, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["id"] if "id" in user.keys() else user.get("id"),
            action,
            user["name"] if "name" in user.keys() else user.get("name"),
            user["email"] if "email" in user.keys() else user.get("email"),
            user["phone"] if "phone" in user.keys() else user.get("phone"),
            user["role"] if "role" in user.keys() else user.get("role"),
            user["profile_pic_path"] if "profile_pic_path" in user.keys() else user.get("profile_pic_path"),
            now_iso(),
        ),
    )


def get_account_form_data() -> dict:
    # Collects and cleans account fields from signup/admin forms in one place.
    school_id = to_int(request.form.get("school_id"))
    school = get_school_by_id(school_id)
    department = request.form.get("department", "").strip()
    if not department:
        department_id = to_int(request.form.get("department_id"))
        department_row = fetch_one("SELECT * FROM departments WHERE id = ?", (department_id,)) if department_id else None
        department = department_row["name"] if department_row else ""
        if department_row and not school:
            school = get_school_by_id(department_row["school_id"])
            school_id = department_row["school_id"]
    return {
        "name": request.form.get("name", "").strip(),
        "email": request.form.get("email", "").strip().lower(),
        "phone": request.form.get("phone", "").strip(),
        "registration_number": request.form.get("registration_number", "").strip().upper(),
        "school_id": school_id,
        "school_name": school["name"] if school else request.form.get("school_name", "").strip(),
        "department": department,
        "study_year": request.form.get("study_year", "").strip(),
    }


def get_user_by_identifier(identifier: str):
    # Lets one login box work for username, email, phone, or registration number.
    normalized = identifier.strip().lower()
    return fetch_one(
        "SELECT * FROM users WHERE lower(username) = ? OR lower(email) = ? OR phone = ? OR upper(registration_number) = ?",
        (normalized, normalized, normalized, identifier.strip().upper()),
    )


def get_user_by_email(email: str):
    # Used by forgot password, which matches users by email only.
    normalized = email.strip().lower()
    return fetch_one(
        "SELECT * FROM users WHERE role != 'admin' AND lower(email) = ?",
        (normalized,),
    )


def get_record_form_data() -> dict:
    # Collects and cleans event record fields from admin forms.
    event_date = request.form.get("event_date", "").strip() or now_ist().date().isoformat()
    academic_year = request.form.get("academic_year", "").strip() or event_date[:4]
    return {
        "student_id": request.form.get("student_id", "").strip(),
        "student_name": request.form.get("student_name", "").strip(),
        "competition_name": request.form.get("competition_name", "").strip(),
        "school_name": request.form.get("school_name", "").strip(),
        "department": request.form.get("department", "").strip(),
        "study_year": request.form.get("study_year", "").strip(),
        "event_id": request.form.get("event_id", "").strip(),
        "event_name": request.form.get("event_name", "").strip(),
        "event_date": event_date,
        "team_id": request.form.get("team_id", "").strip(),
        "team_size": to_int(request.form.get("team_size")) or 0,
        "team_members": request.form.get("team_members", "").strip(),
        "position": request.form.get("position", "").strip(),
        "result": request.form.get("result", "").strip(),
        "prize_money": to_int(request.form.get("prize_money")) or 0,
        "faculty_coordinator": request.form.get("faculty_coordinator", "").strip(),
        "venue": request.form.get("venue", "").strip(),
        "academic_year": academic_year,
    }


# ---------------------------------------------------------------------------
# Public and user-facing routes
# ---------------------------------------------------------------------------

@app.route("/")
def home():
    # Landing page shown before login.
    if g.current_user:
        return redirect(url_for("dashboard"))
    return redirect(url_for("participant_login"))


@app.route("/signup", methods=["GET", "POST"])
def signup():
    # Public signup route for student/participant accounts.
    if request.method == "POST":
        # Public signup always creates a student account.
        role = "student"
        username = request.form.get("username", "").strip().lower()
        form_data = get_account_form_data()
        password = request.form.get("password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()
        if not all([username, form_data["name"], form_data["email"], form_data["phone"], form_data["registration_number"], password, confirm_password]):
            flash("Username, name, email, phone, registration number, password, and confirm password are required.", "error")
            return render_template("signup.html")

        if password != confirm_password:
            flash("Password and confirm password do not match.", "error")
            return render_template("signup.html")

        if len(password) < 6:
            flash("Password must be at least 6 characters.", "error")
            return render_template("signup.html")

        duplicate = fetch_one(
            "SELECT id FROM users WHERE username = ? OR email = ? OR phone = ? OR registration_number = ?",
            (username, form_data["email"], form_data["phone"], form_data["registration_number"]),
        )
        if duplicate:
            flash("An account with that username, email, phone, or registration number already exists.", "error")
            return render_template("signup.html")

        # Save an uploaded profile picture if provided; otherwise use the default avatar.
        profile_pic_path = save_uploaded_file(request.files.get("profile_pic"), "profile") or default_profile_pic()
        execute(
            """
            INSERT INTO users (
                role, username, password_hash, name, email, phone, registration_number,
                school_id, school_name, department, study_year, profile_pic_path, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                role,
                username,
                generate_password_hash(password),
                form_data["name"],
                form_data["email"],
                form_data["phone"],
                form_data["registration_number"],
                form_data["school_id"],
                form_data["school_name"],
                form_data["department"],
                form_data["study_year"],
                profile_pic_path,
                now_iso(),
            ),
        )
        created_user = fetch_one("SELECT * FROM users WHERE username = ?", (username,))
        log_activity("account_created", created_user)
        flash("Sign up complete. You can now log in with email or phone plus password.", "success")
        return redirect(url_for("participant_login"))

    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def participant_login():
    # Shared login page for admin, faculty, coordinators, organisers, and students.
    if request.method == "POST":
        # One identifier field accepts username, email, phone, or registration number.
        identifier = request.form.get("identifier", "").strip()
        password = request.form.get("password", "").strip()
        user = get_user_by_identifier(identifier)

        if not user:
            flash("No account found for that username, email, or phone.", "error")
            return render_template("login.html")

        if not user["password_hash"]:
            flash("This account does not have a password yet. Use forgot password to set one.", "error")
            return render_template("login.html")

        if not check_password_hash(user["password_hash"], password):
            flash("Invalid password.", "error")
            return render_template("login.html")

        # Store the logged-in user id in the session so Flask remembers the user.
        session["user_id"] = user["id"]
        session.pop("pending_user_id", None)
        session.pop("reset_user_id", None)
        update_last_login(user["id"])
        log_activity("login", user)
        flash(f"Welcome, {user['name']}.", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    # Password resets are handled by admin now, so this page only guides users.
    if request.method == "POST":
        flash("Contact admin to change your password.", "error")
        return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html")


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    # Hidden admin URL now redirects to the shared login page.
    return redirect(url_for("participant_login"))


@app.route("/dashboard")
@login_required
def dashboard():
    # Main dashboard after login with summary stats and shortcuts.
    portal_data = build_portal_data()
    return render_template("index.html", **portal_data)


@app.route("/event-center")
@login_required
def event_center():
    # Shows created live events, optionally filtered by school.
    selected_school = request.args.get("school", "").strip()
    events = get_created_events_with_competitions()
    if selected_school:
        events = [event for event in events if (event.get("school_name") or "") == selected_school]
    open_events = [event for event in events if bool(event.get("registration_open", True))]
    closed_events = [event for event in events if not bool(event.get("registration_open", True))]
    return render_template(
        "event_center.html",
        events=open_events,
        open_events=open_events,
        closed_events=closed_events,
        selected_school=selected_school,
    )


@app.route("/event-center/<int:event_id>", methods=["GET", "POST"])
@login_required
def event_detail(event_id: int):
    # Shows one event and handles student registration for that event.
    event = get_event_by_id(event_id)
    if not event:
        abort(404)

    if request.method == "POST":
        action = request.form.get("action", "register")
        if action == "announce_result":
            if g.current_user["role"] != "admin":
                flash("Only admin can announce live event results.", "error")
                return redirect(url_for("event_detail", event_id=event_id))

            competition_id = to_int(request.form.get("announce_competition_id"))
            competition = fetch_one(
                "SELECT * FROM event_competitions WHERE id = ? AND portal_event_id = ?",
                (competition_id, event_id),
            ) if competition_id else None
            participant_name = request.form.get("participant_name", "").strip()
            registration_number = request.form.get("registration_number", "").strip().upper()
            team_name = request.form.get("team_name", "").strip()
            position = request.form.get("position", "").strip()
            result_label = request.form.get("result_label", "").strip()
            prize_money = to_int(request.form.get("announced_prize_money"))
            matched_registration = None
            if registration_number:
                matched_registration = fetch_one(
                    """
                    SELECT
                        r.*,
                        u.registration_number,
                        u.name AS account_name
                    FROM event_registrations r
                    LEFT JOIN users u ON u.id = r.user_id
                    WHERE r.portal_event_id = ? AND upper(u.registration_number) = ?
                    ORDER BY r.created_at DESC
                    LIMIT 1
                    """,
                    (event_id, registration_number),
                )
            if matched_registration:
                participant_name = participant_name or matched_registration.get("participant_name") or matched_registration.get("account_name") or ""
                team_name = team_name or matched_registration.get("team_name") or ""
            if prize_money is None and competition:
                prize_map = {
                    "1": to_int(competition.get("first_prize")) or 0,
                    "2": to_int(competition.get("second_prize")) or 0,
                    "3": to_int(competition.get("third_prize")) or 0,
                }
                prize_money = prize_map.get(position, 0)
            prize_money = prize_money or 0

            if not competition or not participant_name or not position:
                flash("Select competition and enter participant name and position to announce result.", "error")
                return redirect(url_for("event_detail", event_id=event_id))

            execute(
                """
                INSERT INTO event_result_announcements (
                    portal_event_id, competition_id, competition_name, position, result_label,
                    prize_money, team_name, registration_number, participant_name, announced_by, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_id,
                    competition["id"],
                    competition["competition_name"],
                    position,
                    result_label or f"Position {position}",
                    prize_money,
                    team_name,
                    registration_number,
                    participant_name,
                    g.current_user["id"],
                    now_iso(),
                ),
            )
            flash("Result announced for this event.", "success")
            return redirect(url_for("event_detail", event_id=event_id, selected_result_competition=competition["competition_name"]))

        if action == "delete_announcement":
            if g.current_user["role"] != "admin":
                flash("Only admin can delete announced results.", "error")
                return redirect(url_for("event_detail", event_id=event_id))
            announcement_id = to_int(request.form.get("announcement_id"))
            execute(
                "DELETE FROM event_result_announcements WHERE id = ? AND portal_event_id = ?",
                (announcement_id, event_id),
            )
            flash("Announced result removed. Winners data is unchanged.", "success")
            return redirect(url_for("event_detail", event_id=event_id))

        if g.current_user["role"] != "student":
            flash("Only students can register as event participants.", "error")
            return redirect(url_for("event_detail", event_id=event_id))

        if not event["registration_open"]:
            flash("Registration is closed for this event. You can still view the event details.", "error")
            return redirect(url_for("event_detail", event_id=event_id))

        # Prevent the same student from registering twice for the same event.
        existing = fetch_one(
            "SELECT id FROM event_registrations WHERE portal_event_id = ? AND user_id = ?",
            (event_id, g.current_user["id"]),
        )
        if existing:
            flash("You already registered for this event.", "error")
            return redirect(url_for("event_detail", event_id=event_id))

        participant_name = request.form.get("participant_name", "").strip()
        participant_email = request.form.get("participant_email", "").strip().lower()
        participant_phone = request.form.get("participant_phone", "").strip()
        study_year = request.form.get("study_year", "").strip()
        competition_id = to_int(request.form.get("competition_id"))
        team_name = request.form.get("team_name", "").strip()
        registration_team_id = request.form.get("team_id", "").strip()
        team_member_names = [value.strip() for value in request.form.getlist("team_member_name") if value.strip()]
        competition = fetch_one(
            "SELECT * FROM event_competitions WHERE id = ? AND portal_event_id = ?",
            (competition_id, event_id),
        ) if competition_id else None
        if not all([participant_name, participant_email, participant_phone, study_year, competition]):
            flash("Please fill name, email, phone, year, and select a competition.", "error")
            return redirect(url_for("event_detail", event_id=event_id))

        max_team_members = max(1, to_int(competition.get("max_team_members")) or 1)
        if len(team_member_names) > max_team_members:
            flash(f"Only {max_team_members} team member entries are allowed for this competition.", "error")
            return redirect(url_for("event_detail", event_id=event_id))

        notes = request.form.get("notes", "").strip()
        execute(
            """
            INSERT INTO event_registrations (
                portal_event_id, user_id, participant_name, participant_email, participant_phone,
                school_name, department, study_year, competition_id, competition_name, team_name, team_id, team_members, notes, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                g.current_user["id"],
                participant_name,
                participant_email,
                participant_phone,
                g.current_user["school_name"],
                g.current_user["department"],
                study_year,
                competition["id"],
                competition["competition_name"],
                team_name,
                registration_team_id,
                ", ".join(team_member_names),
                notes,
                now_iso(),
            ),
        )
        flash("Event registration submitted successfully.", "success")
        return redirect(url_for("event_detail", event_id=event_id))

    already_registered = False
    if g.current_user["role"] != "admin":
        already_registered = fetch_one(
            "SELECT id FROM event_registrations WHERE portal_event_id = ? AND user_id = ?",
            (event_id, g.current_user["id"]),
        ) is not None
    selected_result_competition = request.args.get("selected_result_competition", "").strip()
    announcement_competitions = sorted({row["competition_name"] for row in event["announcements"] if row.get("competition_name")})
    displayed_announcements = event["announcements"]
    if selected_result_competition:
        displayed_announcements = [
            row for row in displayed_announcements if (row.get("competition_name") or "") == selected_result_competition
        ]
    return render_template(
        "event_detail.html",
        event=event,
        already_registered=already_registered,
        announcement_competitions=announcement_competitions,
        selected_result_competition=selected_result_competition,
        displayed_announcements=displayed_announcements,
    )


@app.route("/admin/events/<int:event_id>/toggle-registration", methods=["POST"])
@admin_required
def toggle_event_registration(event_id: int):
    # Lets admin open or close registration without hiding the event details page.
    event = fetch_one("SELECT * FROM portal_events WHERE id = ?", (event_id,))
    if not event:
        flash("Live event not found.", "error")
        return redirect(url_for("admin_records"))

    is_open = bool(event.get("registration_open", True))
    execute(
        "UPDATE portal_events SET registration_open = ? WHERE id = ?",
        (not is_open, event_id),
    )
    flash(
        "Registration opened for this event." if not is_open else "Registration closed for this event.",
        "success",
    )
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/stats")
@login_required
def stats_view():
    # Displays school-based statistics and ranking data.
    portal_data = build_portal_data()
    return render_template("stats.html", **portal_data)


@app.route("/visuals")
@login_required
def visuals_view():
    # Dedicated graph page built from the same event and winner data as the dashboard.
    portal_data = build_portal_data()
    created_events = get_created_events_with_competitions()
    base_visual_records = portal_data["records"] + get_announcement_rows()
    visual_rows = build_visual_rows(base_visual_records, created_events)
    selected_graph = request.args.get("graph", "events_by_school").strip() or "events_by_school"
    selected_graph_type = request.args.get("graph_type", "auto").strip() or "auto"
    selected_school = request.args.get("school", "").strip()
    selected_year = request.args.get("year", "").strip()
    selected_event = request.args.get("event_name", "").strip()
    selected_competition = request.args.get("competition_name", "").strip()
    visual_event_competition_map: dict[str, list[str]] = {}
    for event in created_events:
        event_name = event.get("title") or ""
        if not event_name:
            continue
        visual_event_competition_map[event_name] = sorted(
            {
                competition.get("competition_name")
                for competition in event.get("competitions", [])
                if competition.get("competition_name")
            }
        )
    for row in visual_rows:
        event_name = row.get("event_name") or ""
        competition_name = row.get("competition_name") or row.get("event_name") or ""
        if not event_name or not competition_name:
            continue
        visual_event_competition_map.setdefault(event_name, [])
        if competition_name not in visual_event_competition_map[event_name]:
            visual_event_competition_map[event_name].append(competition_name)
    for event_name in list(visual_event_competition_map.keys()):
        visual_event_competition_map[event_name] = sorted(visual_event_competition_map[event_name])
    filtered_rows = filter_visual_rows(
        visual_rows,
        selected_school=selected_school,
        selected_year=selected_year,
        selected_event=selected_event,
        selected_competition=selected_competition,
    )
    selected_visual = build_selected_visual(filtered_rows, selected_graph, selected_graph_type)
    graph_options = [
        ("events_by_school", "Events By School"),
        ("department_entries", "Department Entries"),
        ("competition_winners", "Competition Winners"),
        ("winners_by_year", "Winners By Year"),
        ("winner_share_school", "Winner Share By School"),
    ]
    graph_type_options = [
        ("auto", "Default"),
        ("bar", "Bar"),
        ("pie", "Pie"),
        ("line", "Line"),
    ]
    competition_options = sorted(visual_event_competition_map.get(selected_event, [])) if selected_event else sorted(
        {(row.get("competition_name") or row.get("event_name") or "") for row in visual_rows if (row.get("competition_name") or row.get("event_name"))}
    )
    event_names = sorted({row.get("event_name") or "" for row in visual_rows if row.get("event_name")} | {event.get("title") or "" for event in created_events if event.get("title")})
    visual_year_options = sorted(
        {
            str(row.get("academic_year") or "")
            for row in visual_rows
            if row.get("academic_year")
        }
        | {
            str(event.get("event_year") or "")
            for event in created_events
            if event.get("event_year")
        },
        reverse=True,
    )
    return render_template(
        "visuals.html",
        selected_graph=selected_graph,
        selected_graph_type=selected_graph_type,
        selected_school=selected_school,
        selected_year=selected_year,
        selected_event=selected_event,
        selected_competition=selected_competition,
        selected_visual=selected_visual,
        graph_options=graph_options,
        graph_type_options=graph_type_options,
        competition_options=competition_options,
        event_names=event_names,
        visual_event_competition_map=visual_event_competition_map,
        visual_year_options=visual_year_options,
        **portal_data,
    )


@app.route("/winners")
@login_required
def winners_view():
    # Shows winner filters and displays matching historic winners plus announced results.
    portal_data = build_portal_data()
    selected_event = request.args.get("event_name", "").strip()
    selected_competition = request.args.get("competition_name", "").strip()
    selected_year = request.args.get("year", "").strip()
    selected_school = request.args.get("school", "").strip()
    rows = portal_data["records"] + get_announcement_rows()
    created_events = get_created_events_with_competitions()
    ready_to_show = all([selected_event, selected_competition, selected_school, selected_year])
    filtered = []
    if ready_to_show:
        for row in rows:
            if selected_event and row["event_name"] != selected_event:
                continue
            if selected_competition and (row.get("competition_name") or row.get("event_name")) != selected_competition:
                continue
            if selected_year and str(row.get("academic_year") or "") != selected_year:
                continue
            if selected_school and (row.get("school_name") or "") != selected_school:
                continue
            if str(row.get("position") or "").strip():
                filtered.append(row)

    filtered = sorted(
        filtered,
        key=lambda row: (
            str(row.get("event_name") or "").lower(),
            str(row.get("competition_name") or row.get("event_name") or "").lower(),
            str(row.get("academic_year") or ""),
            to_int(row.get("position")) or 999,
            str(row.get("student_name") or "").lower(),
        ),
    )
    event_names = sorted({row["event_name"] for row in rows if row.get("event_name")} | {event["title"] for event in created_events if event.get("title")})
    event_competition_map: dict[str, list[str]] = {}
    for event in created_events:
        event_competition_map[event["title"]] = sorted(
            {
                competition.get("competition_name")
                for competition in event.get("competitions", [])
                if competition.get("competition_name")
            }
        )
    for row in rows:
        event_name = row.get("event_name") or ""
        competition_name = row.get("competition_name") or row.get("event_name") or ""
        if not event_name or not competition_name:
            continue
        event_competition_map.setdefault(event_name, [])
        if competition_name not in event_competition_map[event_name]:
            event_competition_map[event_name].append(competition_name)
    for event_name in list(event_competition_map.keys()):
        event_competition_map[event_name] = sorted(event_competition_map[event_name])
    competition_options = set(event_competition_map.get(selected_event, [])) if selected_event else set()
    year_options = sorted({str(row.get("academic_year") or "") for row in rows if row.get("academic_year")}, reverse=True)
    return render_template(
        "winners.html",
        winners=filtered,
        event_names=event_names,
        competition_options=sorted(competition_options),
        event_competition_map=event_competition_map,
        ready_to_show=ready_to_show,
        year_options=year_options,
        selected_event=selected_event,
        selected_competition=selected_competition,
        selected_year=selected_year,
        selected_school=selected_school,
    )


@app.route("/events")
@login_required
def events():
    # Read-only event listing page with department-based filtering.
    portal_data = build_portal_data()
    department_options = sorted({row["department"] for row in portal_data["records"] if row.get("department")})
    return render_template("events.html", department_options=department_options, **portal_data)


@app.route("/leaderboard")
@login_required
def leaderboard():
    # Dedicated ranking page that reuses the shared statistics payload.
    portal_data = build_portal_data()
    return render_template("leaderboard.html", **portal_data)


@app.route("/students")
@admin_required
def students():
    # Profile browsing page split into student, faculty, and other account groups.
    portal_data = build_portal_data()
    query = request.args.get("q", "").strip()
    school_filter = request.args.get("school", "").strip()
    profile_type = request.args.get("type", "student").strip()
    sql = "SELECT * FROM users WHERE role != 'admin'"
    params: list[str] = []
    if profile_type == "student":
        sql += " AND role = 'student'"
    elif profile_type == "faculty":
        sql += " AND role = 'faculty'"
    elif profile_type == "other":
        sql += " AND role IN ('coordinator', 'event organiser')"
    if query:
        sql += " AND (lower(name) LIKE ? OR phone LIKE ? OR upper(registration_number) LIKE ?)"
        params.extend([f"%{query.lower()}%", f"%{query}%", f"%{query.upper()}%"])
    if school_filter:
        sql += " AND school_name = ?"
        params.append(school_filter)
    sql += " ORDER BY created_at DESC"
    accounts = fetch_all(sql, tuple(params))
    return render_template(
        "students.html",
        accounts=accounts,
        search_query=query,
        selected_school=school_filter,
        profile_type=profile_type,
        **portal_data,
    )


@app.route("/students/<int:user_id>")
@admin_required
def student_profile(user_id: int):
    # Shows one user's profile and their participation/winner history.
    account = fetch_one("SELECT * FROM users WHERE id = ? AND role != 'admin'", (user_id,))
    if not account:
        flash("Student profile not found.", "error")
        return redirect(url_for("students"))
    history = get_user_history(account)
    return render_template("profile.html", viewed_account=account, admin_profile_view=True, **history)


@app.route("/about")
@admin_required
def about():
    # About page stays admin-only in this portal.
    portal_data = build_portal_data()
    return render_template("about.html", **portal_data)


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    # Logged-in user's own profile page for password and picture changes.
    if request.method == "POST":
        action = request.form.get("action", "profile_pic")
        if action == "change_password":
            current_password = request.form.get("current_password", "").strip()
            new_password = request.form.get("new_password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()
            if not check_password_hash(g.current_user["password_hash"], current_password):
                flash("Current password is wrong. Use forgot password if needed.", "error")
                return redirect(url_for("profile"))
            if not new_password or not confirm_password:
                flash("Enter the new password and confirm it.", "error")
                return redirect(url_for("profile"))
            if new_password != confirm_password:
                flash("New password and confirm password do not match.", "error")
                return redirect(url_for("profile"))
            execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), g.current_user["id"]))
            flash("Password changed successfully.", "success")
            return redirect(url_for("profile"))

        profile_pic_path = save_uploaded_file(request.files.get("profile_pic"), "profile")
        if profile_pic_path:
            execute("UPDATE users SET profile_pic_path = ? WHERE id = ?", (profile_pic_path, g.current_user["id"]))
            flash("Profile picture updated.", "success")
            return redirect(url_for("profile"))
        flash("Please choose an image first.", "error")
    history = get_user_history(dict(g.current_user))
    return render_template("profile.html", **history)


# ---------------------------------------------------------------------------
# Event manager and admin operations
# ---------------------------------------------------------------------------

@app.route("/admin/records", methods=["GET", "POST"])
@event_manager_required
def admin_records():
    # Central Event Manager screen for live events, records, branding, and exports.
    if request.method == "POST":
        # Posting directly on this route adds a new historic event record/winner row.
        if g.current_user["role"] not in {"admin", "faculty", "event organiser"}:
            flash("Only admin, faculty, and event organisers can add event entries.", "error")
            return redirect(url_for("admin_records"))
        data = get_record_form_data()
        if not data["student_name"] or not data["event_name"]:
            flash("Student name and event name are required.", "error")
        else:
            execute(
                """
                INSERT INTO event_records (
                    student_id, student_name, competition_name, school_name, department, study_year, event_id, event_name, event_date,
                    team_id, team_size, team_members, position, result, prize_money,
                    faculty_coordinator, venue, academic_year, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    data["student_id"],
                    data["student_name"],
                    data["competition_name"],
                    data["school_name"],
                    data["department"],
                    data["study_year"],
                    data["event_id"],
                    data["event_name"],
                    data["event_date"],
                    data["team_id"],
                    data["team_size"],
                    data["team_members"],
                    data["position"],
                    data["result"],
                    data["prize_money"],
                    data["faculty_coordinator"],
                    data["venue"],
                    data["academic_year"],
                    g.current_user["id"],
                    now_iso(),
                ),
            )
            flash("Event data added successfully.", "success")
            return redirect(url_for("admin_records"))

    portal_data = build_portal_data()
    created_events = get_created_events_with_competitions()
    portal_settings = fetch_one("SELECT * FROM portal_settings WHERE id = 1")
    portal_data["created_events"] = created_events
    return render_template("admin_records.html", portal_settings=portal_settings, **portal_data)


@app.route("/admin/events/create", methods=["POST"])
@manage_events_required
def create_portal_event():
    # Creates one live event and then saves each of its competition rows.
    # The main event goes into portal_events, while each competition is stored separately.
    title = request.form.get("title", "").strip()
    total_prize_money = to_int(request.form.get("total_prize_money")) or 0
    event_year = request.form.get("event_year", "").strip()
    description = request.form.get("description", "").strip()
    school_id = to_int(request.form.get("school_id"))
    school = get_school_by_id(school_id)
    department_name = request.form.get("department_name", "").strip()
    wallpaper_path = save_uploaded_file(request.files.get("event_wallpaper"), "event_wallpaper")
    poster_image_path = save_uploaded_file(request.files.get("event_poster"), "event_poster")
    competition_indexes = sorted(
        {
            int(key.rsplit("_", 1)[1])
            for key in request.form
            if key.startswith("competition_name_") and key.rsplit("_", 1)[1].isdigit()
        }
    )
    competition_count = len([index for index in competition_indexes if request.form.get(f"competition_name_{index}", "").strip()])

    if not title:
        flash("Event title is required.", "error")
        return redirect(url_for("admin_records"))

    event_id = execute(
        """
        INSERT INTO portal_events (
            title, total_prize_money, competition_count, event_year, school_id, school_name, department_name,
            description, wallpaper_path, poster_image_path, created_by, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            title,
            total_prize_money,
            competition_count,
            event_year,
            school_id,
            school["name"] if school else "",
            department_name,
            description,
            wallpaper_path,
            poster_image_path,
            g.current_user["id"],
            now_iso(),
        ),
    )

    created = 0
    for index in competition_indexes:
        competition_name = request.form.get(f"competition_name_{index}", "").strip()
        if not competition_name:
            continue
        execute(
            """
            INSERT INTO event_competitions (portal_event_id, competition_name, venue, max_team_members, first_prize, second_prize, third_prize, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                competition_name,
                request.form.get(f"competition_venue_{index}", "").strip(),
                max(1, to_int(request.form.get(f"max_team_members_{index}")) or 1),
                to_int(request.form.get(f"first_prize_{index}")) or 0,
                to_int(request.form.get(f"second_prize_{index}")) or 0,
                to_int(request.form.get(f"third_prize_{index}")) or 0,
                now_iso(),
            ),
        )
        created += 1

    flash(f"New event created with {created} competition entries.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/events/<int:event_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_portal_event(event_id: int):
    # Lets admin update the main event information and its existing competition rows.
    event = get_event_by_id(event_id)
    if not event:
        flash("Live event not found.", "error")
        return redirect(url_for("admin_records"))

    if request.method == "POST":
        school_id = to_int(request.form.get("school_id"))
        school = get_school_by_id(school_id)
        wallpaper_path = save_uploaded_file(request.files.get("event_wallpaper"), "event_wallpaper")
        poster_image_path = save_uploaded_file(request.files.get("event_poster"), "event_poster")
        execute(
            """
            UPDATE portal_events
            SET title = ?, total_prize_money = ?, event_year = ?, school_id = ?, school_name = ?, department_name = ?,
                description = ?, wallpaper_path = ?, poster_image_path = ?, registration_open = ?
            WHERE id = ?
            """,
            (
                request.form.get("title", "").strip() or event["title"],
                to_int(request.form.get("total_prize_money")) or 0,
                request.form.get("event_year", "").strip(),
                school_id,
                school["name"] if school else event.get("school_name", ""),
                request.form.get("department_name", "").strip(),
                request.form.get("description", "").strip(),
                wallpaper_path or event.get("wallpaper_path"),
                poster_image_path or event.get("poster_image_path"),
                request.form.get("registration_open") == "on",
                event_id,
            ),
        )
        for competition in event["competitions"]:
            prefix = f"competition_{competition['id']}_"
            execute(
                """
                UPDATE event_competitions
                SET competition_name = ?, venue = ?, max_team_members = ?, first_prize = ?, second_prize = ?, third_prize = ?
                WHERE id = ? AND portal_event_id = ?
                """,
                (
                    request.form.get(f"{prefix}name", "").strip() or competition["competition_name"],
                    request.form.get(f"{prefix}venue", "").strip(),
                    max(1, to_int(request.form.get(f"{prefix}max_team_members")) or 1),
                    to_int(request.form.get(f"{prefix}first_prize")) or 0,
                    to_int(request.form.get(f"{prefix}second_prize")) or 0,
                    to_int(request.form.get(f"{prefix}third_prize")) or 0,
                    competition["id"],
                    event_id,
                ),
            )
        flash("Live event updated.", "success")
        return redirect(url_for("admin_records"))

    portal_data = build_portal_data()
    return render_template("edit_event.html", event=event, **portal_data)


@app.route("/admin/events/<int:event_id>/delete", methods=["POST"])
@admin_required
def delete_portal_event(event_id: int):
    # Admin-only hard delete for a live event and its related rows.
    event = fetch_one("SELECT * FROM portal_events WHERE id = ?", (event_id,))
    if not event:
        flash("Live event not found.", "error")
        return redirect(url_for("admin_records"))

    execute("DELETE FROM event_registrations WHERE portal_event_id = ?", (event_id,))
    execute("DELETE FROM event_competitions WHERE portal_event_id = ?", (event_id,))
    execute("DELETE FROM portal_events WHERE id = ?", (event_id,))
    flash(f"{event['title']} deleted from live events.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/branding", methods=["POST"])
@admin_required
def update_branding():
    # Updates the website name, logo, and per-page wallpaper settings.
    site_name = request.form.get("site_name", "").strip()
    logo_path = save_uploaded_file(request.files.get("logo_image"), "logo")
    wallpaper_path = save_uploaded_file(request.files.get("wallpaper_image"), "wallpaper")
    page_location = request.form.get("page_location", "").strip()
    page_wallpaper_path = save_uploaded_file(request.files.get("page_wallpaper_image"), "page_wallpaper")
    current = fetch_one("SELECT * FROM portal_settings WHERE id = 1")
    allowed_page_columns = {
        "login": "login_wallpaper_path",
        "dashboard": "dashboard_wallpaper_path",
        "events": "event_wallpaper_path",
        "stats": "stats_wallpaper_path",
        "winners": "winners_wallpaper_path",
        "admin": "admin_wallpaper_path",
    }
    execute(
        """
        UPDATE portal_settings
        SET site_name = ?, logo_path = ?, wallpaper_path = ?, updated_at = ?
        WHERE id = 1
        """,
        (
            site_name or current["site_name"] or "College Event Portal",
            logo_path or current["logo_path"],
            wallpaper_path or current["wallpaper_path"],
            now_iso(),
        ),
    )
    if page_wallpaper_path and page_location in allowed_page_columns:
        execute(
            f"UPDATE portal_settings SET {allowed_page_columns[page_location]} = ?, updated_at = ? WHERE id = 1",
            (page_wallpaper_path, now_iso()),
        )
    flash("Portal branding updated.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/records/<int:record_id>/edit", methods=["GET", "POST"])
@manage_events_required
def edit_record(record_id: int):
    # Edits one saved historic/event winner record.
    record = fetch_one("SELECT * FROM event_records WHERE id = ?", (record_id,))
    if not record:
        flash("Record not found.", "error")
        return redirect(url_for("admin_records"))

    if request.method == "POST":
        data = get_record_form_data()
        execute(
            """
            UPDATE event_records
            SET student_id = ?, student_name = ?, competition_name = ?, school_name = ?, department = ?, study_year = ?, event_id = ?, event_name = ?,
                event_date = ?, team_id = ?, team_size = ?, team_members = ?, position = ?, result = ?,
                prize_money = ?, faculty_coordinator = ?, venue = ?, academic_year = ?
            WHERE id = ?
            """,
            (
                data["student_id"],
                data["student_name"],
                data["competition_name"],
                data["school_name"],
                data["department"],
                data["study_year"],
                data["event_id"],
                data["event_name"],
                data["event_date"],
                data["team_id"],
                data["team_size"],
                data["team_members"],
                data["position"],
                data["result"],
                data["prize_money"],
                data["faculty_coordinator"],
                data["venue"],
                data["academic_year"],
                record_id,
            ),
        )
        flash("Record updated successfully.", "success")
        return redirect(url_for("admin_records"))

    return render_template("edit_record.html", record=record)


@app.route("/admin/records/<int:record_id>/delete", methods=["POST"])
@admin_required
def delete_record(record_id: int):
    # Admin-only delete for one event record row.
    execute("DELETE FROM event_records WHERE id = ?", (record_id,))
    flash("Record deleted.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/accounts")
@admin_required
def admin_accounts():
    # Admin overview of all accounts plus recent activity history.
    accounts = fetch_all("SELECT * FROM users ORDER BY CASE WHEN role = 'admin' THEN 0 ELSE 1 END, created_at DESC")
    activity_logs = fetch_all("SELECT * FROM activity_logs ORDER BY created_at DESC LIMIT 30")
    year_groups = fetch_all(
        """
        SELECT COALESCE(NULLIF(study_year, ''), 'Not set') AS year_label, COUNT(*) AS total
        FROM users
        WHERE role != 'admin'
        GROUP BY year_label
        ORDER BY year_label
        """
    )
    return render_template(
        "admin_accounts.html",
        accounts=accounts,
        activity_logs=activity_logs,
        year_groups=year_groups,
        schools=get_schools(),
        departments=get_departments(),
    )


@app.route("/admin/accounts/create", methods=["POST"])
@admin_required
def create_account():
    # Admin-created account flow for student and staff roles.
    role = request.form.get("role", "").strip().lower()
    username = request.form.get("username", "").strip().lower()
    form_data = get_account_form_data()
    password = request.form.get("password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()

    if role not in {"student", "faculty", "coordinator", "event organiser"}:
        flash("Admin can create only student, faculty, coordinator, or event organiser accounts.", "error")
        return redirect(url_for("admin_accounts"))

    if not all([username, form_data["name"], form_data["email"], form_data["phone"], password, confirm_password]):
        flash("Role, username, name, email, phone, password, and confirm password are required.", "error")
        return redirect(url_for("admin_accounts"))

    if password != confirm_password:
        flash("Password and confirm password do not match.", "error")
        return redirect(url_for("admin_accounts"))

    duplicate = fetch_one(
        "SELECT id FROM users WHERE username = ? OR email = ? OR phone = ? OR registration_number = ?",
        (username, form_data["email"], form_data["phone"], form_data["registration_number"] or None),
    )
    if duplicate:
        flash("An account with that username, email, phone, or registration number already exists.", "error")
        return redirect(url_for("admin_accounts"))

    execute(
        """
        INSERT INTO users (
            role, username, password_hash, name, email, phone, registration_number,
            school_id, school_name, department, study_year, profile_pic_path, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            role,
            username,
            generate_password_hash(password),
            form_data["name"],
            form_data["email"],
            form_data["phone"],
            form_data["registration_number"] or None,
            form_data["school_id"],
            form_data["school_name"],
            form_data["department"],
            form_data["study_year"],
            default_profile_pic(),
            now_iso(),
        ),
    )
    created_user = fetch_one("SELECT * FROM users WHERE username = ?", (username,))
    log_activity("account_created_by_admin", created_user)
    flash(f"{role.title()} account created successfully.", "success")
    return redirect(url_for("admin_accounts"))


@app.route("/admin/accounts/<int:user_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_account(user_id: int):
    # Lets admin correct stored account details such as phone, reg no, school, and department.
    account = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,))
    if not account:
        flash("Account not found.", "error")
        return redirect(url_for("admin_accounts"))

    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        form_data = get_account_form_data()
        role = request.form.get("role", "").strip().lower()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()
        if account["role"] == "admin":
            role = "admin"

        duplicate = fetch_one(
            "SELECT id FROM users WHERE id != ? AND (username = ? OR email = ? OR phone = ? OR registration_number = ?)",
            (user_id, username, form_data["email"], form_data["phone"], form_data["registration_number"] or None),
        )
        if duplicate:
            flash("Another account already uses that username, email, phone, or registration number.", "error")
            return render_template("edit_account.html", account=account, schools=get_schools(), departments=get_departments())

        if new_password or confirm_password:
            if new_password != confirm_password:
                flash("New password and confirm password do not match.", "error")
                return render_template("edit_account.html", account=account, schools=get_schools(), departments=get_departments())
            if len(new_password) < 6:
                flash("New password must be at least 6 characters.", "error")
                return render_template("edit_account.html", account=account, schools=get_schools(), departments=get_departments())

        execute(
            """
            UPDATE users
            SET role = ?, username = ?, name = ?, email = ?, phone = ?, registration_number = ?, school_id = ?, school_name = ?, department = ?, study_year = ?
            WHERE id = ?
            """,
            (
                role,
                username,
                form_data["name"],
                form_data["email"],
                form_data["phone"],
                form_data["registration_number"] or None,
                form_data["school_id"],
                form_data["school_name"],
                form_data["department"],
                form_data["study_year"],
                user_id,
            ),
        )
        if new_password:
            execute(
                "UPDATE users SET password_hash = ?, otp_code = NULL, otp_created_at = NULL WHERE id = ?",
                (generate_password_hash(new_password), user_id),
            )
        flash("Account updated successfully.", "success")
        return redirect(url_for("admin_accounts"))

    return render_template("edit_account.html", account=account, schools=get_schools(), departments=get_departments())


@app.route("/admin/accounts/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_account(user_id: int):
    # Deletes a non-admin account from the system.
    account = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,))
    if not account:
        flash("Account not found.", "error")
    elif account["role"] == "admin":
        flash("Admin account cannot be deleted.", "error")
    else:
        execute("DELETE FROM users WHERE id = ?", (user_id,))
        flash("Account deleted.", "success")
    return redirect(url_for("admin_accounts"))


@app.route("/admin/schools/create", methods=["POST"])
@admin_required
def create_school():
    # Adds a school and can also create its first department in the same step.
    name = request.form.get("school_name", "").strip()
    first_department = request.form.get("department_name", "").strip()
    if not name:
        flash("School name is required.", "error")
    elif fetch_one("SELECT id FROM schools WHERE lower(name) = lower(?)", (name,)):
        flash("That school already exists.", "error")
    else:
        school_id = execute("INSERT INTO schools (name, created_at) VALUES (?, ?)", (name, now_iso()))
        if first_department:
            execute(
                "INSERT INTO departments (school_id, name, created_at) VALUES (?, ?, ?)",
                (school_id, first_department, now_iso()),
            )
            flash("School and first department added.", "success")
        else:
            flash("School added.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/schools/<int:school_id>/edit", methods=["POST"])
@admin_required
def edit_school(school_id: int):
    # Renames a school and updates linked tables that store the school name.
    name = request.form.get("school_name", "").strip()
    school = fetch_one("SELECT * FROM schools WHERE id = ?", (school_id,))
    if not school:
        flash("School not found.", "error")
    elif not name:
        flash("School name is required.", "error")
    else:
        execute("UPDATE schools SET name = ? WHERE id = ?", (name, school_id))
        execute("UPDATE users SET school_name = ? WHERE school_id = ?", (name, school_id))
        execute("UPDATE event_records SET school_name = ? WHERE school_name = ?", (name, school["name"]))
        execute("UPDATE portal_events SET school_name = ? WHERE school_id = ?", (name, school_id))
        execute("UPDATE event_registrations SET school_name = ? WHERE school_name = ?", (name, school["name"]))
        flash("School updated.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/schools/<int:school_id>/delete", methods=["POST"])
@admin_required
def delete_school(school_id: int):
    # Deletes a school only when no departments remain under it.
    if fetch_one("SELECT id FROM departments WHERE school_id = ?", (school_id,)):
        flash("Delete or move the departments in that school first.", "error")
    else:
        execute("DELETE FROM schools WHERE id = ?", (school_id,))
        flash("School deleted.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/departments/create", methods=["POST"])
@admin_required
def create_department():
    # Adds a department under the selected school.
    school_id = to_int(request.form.get("school_id"))
    name = request.form.get("department_name", "").strip()
    if not school_id or not name:
        flash("School and department name are required.", "error")
    elif fetch_one("SELECT id FROM departments WHERE school_id = ? AND lower(name) = lower(?)", (school_id, name)):
        flash("That department already exists in the selected school.", "error")
    else:
        execute("INSERT INTO departments (school_id, name, created_at) VALUES (?, ?, ?)", (school_id, name, now_iso()))
        flash("Department added.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/departments/<int:department_id>/edit", methods=["POST"])
@admin_required
def edit_department(department_id: int):
    # Renames a department and updates related rows that still use the old name.
    name = request.form.get("department_name", "").strip()
    department = fetch_one("SELECT * FROM departments WHERE id = ?", (department_id,))
    if not department:
        flash("Department not found.", "error")
    elif not name:
        flash("Department name is required.", "error")
    else:
        execute("UPDATE departments SET name = ? WHERE id = ?", (name, department_id))
        execute("UPDATE users SET department = ? WHERE department = ?", (name, department["name"]))
        execute("UPDATE event_records SET department = ? WHERE department = ?", (name, department["name"]))
        execute("UPDATE portal_events SET department_name = ? WHERE department_name = ?", (name, department["name"]))
        execute("UPDATE event_registrations SET department = ? WHERE department = ?", (name, department["name"]))
        flash("Department updated.", "success")
    return redirect(url_for("admin_records"))


@app.route("/admin/departments/<int:department_id>/delete", methods=["POST"])
@admin_required
def delete_department(department_id: int):
    # Removes a department from the school/department manager.
    execute("DELETE FROM departments WHERE id = ?", (department_id,))
    flash("Department deleted.", "success")
    return redirect(url_for("admin_records"))


# ---------------------------------------------------------------------------
# CSV download/export routes
# ---------------------------------------------------------------------------

@app.route("/admin/export")
@export_required
def export_records():
    # Exports winner/history records as CSV, optionally filtered by year or school.
    year = request.args.get("year", "").strip()
    school = request.args.get("school", "").strip()
    conditions = []
    params: list[str] = []
    if year:
        conditions.append("academic_year = ?")
        params.append(year)
    if school:
        conditions.append("school_name = ?")
        params.append(school)
    where_clause = f" WHERE {' AND '.join(conditions)}" if conditions else ""
    rows = fetch_all(f"SELECT * FROM event_records{where_clause} ORDER BY event_date DESC, id DESC", tuple(params))
    filename = "event_records"
    if year:
        filename += f"_{year}"
    if school:
        filename += f"_{school.replace(' ', '_').lower()}"
    filename += ".csv"

    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(
        [
            "student_id",
            "student_name",
            "competition_name",
            "school_name",
            "department",
            "study_year",
            "event_id",
            "event_name",
            "event_date",
            "team_id",
            "team_size",
            "team_members",
            "position",
            "result",
            "prize_money",
            "faculty_coordinator",
            "venue",
            "academic_year",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["student_id"],
                row["student_name"],
                row["competition_name"],
                row["school_name"],
                row["department"],
                row["study_year"],
                row["event_id"],
                row["event_name"],
                row["event_date"],
                row["team_id"],
                row["team_size"],
                row["team_members"],
                row["position"],
                row["result"],
                row["prize_money"],
                row["faculty_coordinator"],
                row["venue"],
                row["academic_year"],
            ]
        )

    output = io.BytesIO(csv_buffer.getvalue().encode("utf-8"))
    output.seek(0)
    return send_file(output, mimetype="text/csv", as_attachment=True, download_name=filename)


@app.route("/download/events")
@login_required
def download_created_events():
    # Downloads the live event setup list as CSV.
    school = request.args.get("school", "").strip()
    rows = get_created_events_with_competitions()
    if school:
        rows = [row for row in rows if (row.get("school_name") or "") == school]
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["event_title", "school_name", "department_name", "event_year", "total_prize_money", "competition_count", "description"])
    for row in rows:
        writer.writerow([row["title"], row.get("school_name"), row.get("department_name"), row["event_year"], row["total_prize_money"], row["competition_count"], row["description"]])
    output = io.BytesIO(csv_buffer.getvalue().encode("utf-8"))
    output.seek(0)
    return send_file(output, mimetype="text/csv", as_attachment=True, download_name="created_events.csv")


@app.route("/download/stats")
@login_required
def download_stats():
    # Downloads the school ranking table as CSV.
    portal_data = build_portal_data()
    school = request.args.get("school", "").strip()
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["school", "points", "wins", "participants", "prize_money"])
    rankings = portal_data["department_rankings"]
    if school:
        rankings = [row for row in rankings if row["school"] == school]
    for row in rankings:
        writer.writerow([row["school"], row["points"], row["wins"], row["participant_count"], row["prize_money"]])
    output = io.BytesIO(csv_buffer.getvalue().encode("utf-8"))
    output.seek(0)
    return send_file(output, mimetype="text/csv", as_attachment=True, download_name="stats.csv")


@app.route("/download/winners")
@login_required
def download_winners():
    # Downloads winner data using the same filters as the winners page.
    event_name = request.args.get("event_name", "").strip()
    competition_name = request.args.get("competition_name", "").strip()
    year = request.args.get("year", "").strip()
    school = request.args.get("school", "").strip()
    portal_data = build_portal_data()
    rows = []
    for row in portal_data["records"]:
        if event_name and row["event_name"] != event_name:
            continue
        if competition_name and (row.get("competition_name") or row.get("event_name")) != competition_name:
            continue
        if year and str(row.get("academic_year") or "") != year:
            continue
        if school and (row.get("school_name") or "") != school:
            continue
        if str(row.get("position") or "").strip() in {"1", "2", "3"}:
            rows.append(row)
    rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("event_name") or "").lower(),
            str(row.get("competition_name") or row.get("event_name") or "").lower(),
            str(row.get("academic_year") or ""),
            to_int(row.get("position")) or 999,
            str(row.get("student_name") or "").lower(),
        ),
    )
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(["event_name", "competition_name", "school_name", "student_name", "department", "position", "result", "prize_money", "academic_year"])
    for row in rows:
        writer.writerow([row["event_name"], row.get("competition_name") or row.get("event_name"), row.get("school_name"), row["student_name"], row["department"], row["position"], row["result"], row["prize_money"], row["academic_year"]])
    output = io.BytesIO(csv_buffer.getvalue().encode("utf-8"))
    output.seek(0)
    return send_file(output, mimetype="text/csv", as_attachment=True, download_name="winners.csv")


@app.route("/logout")
def logout():
    # Clears the session and records the logout activity.
    if getattr(g, "current_user", None):
        log_activity("logout", g.current_user)
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("home"))


# Load local email and Supabase settings before touching the database.
load_email_settings()
load_supabase_settings()


with app.app_context():
    # Initialize tables, defaults, and migrations once when the app starts.
    init_portal()


if __name__ == "__main__":
    # Local development entry point.
    app.run(debug=True)
